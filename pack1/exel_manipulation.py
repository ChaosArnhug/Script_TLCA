from openpyxl import load_workbook, Workbook


def _copy_data(path_file_data: str, sheet_name: str) -> list:
    """
    Function that make a copy of the data within the exel sample
    :param path_file_data: the path of the exel sample
    :param sheet_name: the exel'sheet where the data is
    :return: A list of a list which contain the data per row
    """
    wb = load_workbook(filename=path_file_data)
    ws = wb[sheet_name]
    all_rows = []
    for row in range(1, len(ws["A"]) + 1):
        curr_row = []
        for column in range(1, len(ws["1"]) + 1):
            curr_row.append(ws.cell(row=row, column=column).value)
        all_rows.append(curr_row)
    wb.save(path_file_data)
    return all_rows


def _construct_data(list_data: list, student_data: list, actives_abilities: list) -> list:
    """
    Function that complete the unfilled data from the sample
    :param list_data: the sample's data
    :param student_data: the list which contains the data of each student
    :param actives_abilities: Determine which abilities are actives for the test
    :return: A list of a list which contain the complete data per row
    """
    first_row = True
    for students in range(len(list_data)):
        if first_row:
            first_row = False
            continue
        if student_data[students - 1].is_finished:
            score = round(student_data[students - 1].result)
            list_data[students][4] = "x"  # Select
            list_data[students][6] = str(score) + "%"   # Comment
            for ability_nb in range(len(actives_abilities)):
                if actives_abilities[ability_nb] == "True":
                    list_data[students][7 + ability_nb] = "x"
        else:
            list_data[students][4] = ""     # Select
            list_data[students][6] = ""     # Comment
            for ability_nb in range(len(actives_abilities)):
                list_data[students][7 + ability_nb] = ""

        list_data[students][5] = student_data[students - 1].date    # Date
        students += 1
    return list_data


def _write_data(list_data: list, file_name: str, sheet_name: str):
    """
    Function that write on a new exel the complete data from the evaluation
    :param list_data: The complete data
    :param file_name: The exel file's name
    :param sheet_name:The sheet's name within the exel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in range(1, len(list_data)):
        for column in range(1, len(list_data[0])):
            ws.cell(row=row, column=column).value = list_data[row - 1][column]
    wb.save(file_name)


def manage_exel(path_file_data: str, sheet_name_origin: str, student_data: list, actives_abilities: list,
                file_name_end: str, sheet_name_end: str):
    """
    Execute the copy of the data, the completion of it and the writing into a new exel
    :param path_file_data:the path of the exel sample
    :param sheet_name_origin: the exel'sheet where the data is
    :param student_data: the list which contains the data of each student
    :param actives_abilities::param actives_abilities: Determine which abilities are actives for the test
    :param file_name_end: The exel file's name which contain the final data
    :param sheet_name_end: The sheet's name within the exel file
    """
    _write_data(_construct_data(_copy_data(path_file_data, sheet_name_origin), student_data, actives_abilities),
                file_name_end, sheet_name_end)
